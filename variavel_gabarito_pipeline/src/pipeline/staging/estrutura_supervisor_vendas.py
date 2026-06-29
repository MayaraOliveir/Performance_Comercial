"""Staging da Estrutura Supervisor Vendas materializada no gabarito."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from pipeline.core.logger import configure_logger
from pipeline.core.paths import load_paths_config
from pipeline.io.csv_writer import write_csv
from pipeline.staging.common import clean_text, create_chave_centro_rota, first_excel_file, to_number


SHEET_NAME = "Estrutura Supervisor Vendas"
SOURCE_TABLE = "Supervisor_Vendas"
META_COLUMNS = {
    "MKO": ("META", "REAL"),
    "MKR": ("META3", "REAL4"),
    "MKD": ("META7", "REAL8"),
    "MCN": ("META2", "REAL3"),
    "RED": ("META4", "REAL5"),
    "FRD": ("META6", "REAL7"),
    "EFV": ("META8", "REAL9"),
}


def transform_estrutura_supervisor_vendas(raw: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Padroniza a estrutura materializada de Supervisor de Vendas."""

    result = raw.copy()
    result = result.dropna(how="all").reset_index(drop=True)
    result = result[result["Centro"].notna() & result["Supervisão"].notna()].copy()

    result = result.rename(
        columns={
            "Gerência": "Gerencia",
            "Supervisão": "Supervisao",
            "Matrícula": "RotaSup",
        }
    )
    for column in ["UF", "Gerencia", "Supervisao", "Centro", "RotaSup", "Nome", "Cargo", "CHAVE2", "CHAVE"]:
        if column in result.columns:
            result[column] = result[column].map(clean_text)

    result["ChaveSupervisor"] = result["Centro"].fillna("") + result["Supervisao"].fillna("")
    result["ChaveRotaSup"] = [
        create_chave_centro_rota(centro, rota)
        for centro, rota in zip(result["Centro"], result["RotaSup"], strict=False)
    ]
    result["TipoRotaSup"] = result["RotaSup"].map(lambda value: value[:2] if isinstance(value, str) else None)
    result["TipoSupervisor"] = result["Supervisao"].map(_tipo_supervisor)
    result["FonteEstrutura"] = "Estrutura Supervisor Vendas materializada no gabarito"
    result["TabelaOrigem"] = SOURCE_TABLE

    for meta_column, real_column in META_COLUMNS.values():
        if meta_column in result.columns:
            result[meta_column] = result[meta_column].map(to_number)
        if real_column in result.columns:
            result[real_column] = result[real_column].map(to_number)

    result = result.rename(columns=_metric_rename_map())
    duplicate_counts = result.groupby("ChaveSupervisor")["ChaveSupervisor"].transform("size")
    result["QtdDuplicidadeChaveSupervisor"] = duplicate_counts
    result["StatusDuplicidadeChaveSupervisor"] = duplicate_counts.map(
        lambda value: "Duplicada na estrutura" if int(value) > 1 else "Unica"
    )
    duplicates = result[result["QtdDuplicidadeChaveSupervisor"] > 1].copy()

    result = (
        result.sort_values(["ChaveSupervisor", "ChaveRotaSup"], na_position="last")
        .drop_duplicates(subset=["ChaveSupervisor"], keep="first")
        .reset_index(drop=True)
    )
    return result, _build_diagnostics(raw, result, duplicates)


def build_staging_estrutura_supervisor_vendas(paths_config_path: str | Path | None = None) -> list[Path]:
    """Le a estrutura no gabarito e salva o staging da figura."""

    paths_config = load_paths_config(paths_config_path) if paths_config_path else load_paths_config()
    logger = configure_logger("pipeline.staging.estrutura_supervisor_vendas")
    input_file = first_excel_file(paths_config.values["raw_sources"], "gabarito_validacao")
    raw = read_estrutura_supervisor_vendas(input_file)
    transformed, diagnostics = transform_estrutura_supervisor_vendas(raw)

    staging_dir = paths_config.values["processed"]["staging"]
    validation_dir = paths_config.values["processed"]["validation"]
    outputs = [
        write_csv(transformed, staging_dir / "stg_estrutura_supervisor_vendas.csv"),
        write_csv(diagnostics, validation_dir / "resumo_estrutura_supervisor_vendas.csv"),
    ]
    logger.info(
        "staging_estrutura_supervisor_vendas_concluido",
        extra={"arquivo": str(input_file), "linhas_lidas": len(raw), "linhas_salvas": len(transformed)},
    )
    return outputs


def read_estrutura_supervisor_vendas(file_path: str | Path) -> pd.DataFrame:
    """Le a tabela Supervisor_Vendas pela area conhecida B:Y."""

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    worksheet = workbook[SHEET_NAME]
    rows = list(worksheet.iter_rows(min_row=3, max_row=426, min_col=2, max_col=25, values_only=True))
    workbook.close()
    header = list(rows[0])
    return pd.DataFrame(rows[1:], columns=header)


def _metric_rename_map() -> dict[str, str]:
    rename: dict[str, str] = {}
    for kpi, (meta_column, real_column) in META_COLUMNS.items():
        rename[meta_column] = f"Meta{kpi}Estrutura"
        rename[real_column] = f"Realizado{kpi}Estrutura"
    return rename


def _tipo_supervisor(supervisao: object) -> str:
    text = clean_text(supervisao) or ""
    if "AS 5+" in text:
        return "AS 5+"
    if "AS5+" in text:
        return "AS5+"
    if "SEGMENTAD" in text:
        return "SEGMENTAD"
    return ""


def _build_diagnostics(raw: pd.DataFrame, transformed: pd.DataFrame, duplicates: pd.DataFrame) -> pd.DataFrame:
    rows = [
        {"Metrica": "linhas_raw", "Valor": len(raw), "Categoria": pd.NA},
        {"Metrica": "linhas_staging", "Valor": len(transformed), "Categoria": pd.NA},
        {
            "Metrica": "chaves_supervisor_unicas",
            "Valor": transformed["ChaveSupervisor"].nunique(),
            "Categoria": pd.NA,
        },
        {"Metrica": "linhas_com_chave_duplicada", "Valor": len(duplicates), "Categoria": pd.NA},
    ]
    for status, count in transformed["StatusDuplicidadeChaveSupervisor"].value_counts(dropna=False).items():
        rows.append({"Metrica": "status_duplicidade", "Valor": int(count), "Categoria": status})
    return pd.DataFrame(rows)
